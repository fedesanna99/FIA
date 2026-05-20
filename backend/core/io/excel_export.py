"""
Export FEAModel + risultati in Excel multi-sheet (openpyxl).

Layout sheets:
    - Summary       : nome, descrizione, conteggi, unità
    - Nodes         : id, x, y, z, label
    - Elements      : id, type, n1, n2, ..., material_id, section_id, winkler_k
    - Loads         : id, type, target_id, fx/fy/fz/qy/mass/...
    - Constraints   : id, type, node_id, dofs, spring_k
    - Displacements : node_id, ux, uy, uz, rx, ry, rz   (se results forniti)
    - Reactions     : node_id, fx, fy, fz, mx, my, mz   (se results forniti)
    - ElementForces : element_id, N_i, Vy_i, ..., N_j, ...
    - Modes         : mode, freq_Hz, period_s, participation_x/y/z   (se modal)

Tutti i numeri usano number_format = '0.0000E+00' o '#,##0.000' a seconda
della grandezza.
"""
from __future__ import annotations
from pathlib import Path
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from schemas import FEAModel
from schemas.results import StaticResults, ModalResults


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="305496")
_HEADER_ALIGN = Alignment(horizontal="center", vertical="center")


def _set_header(ws, row: int, headers: list[str]) -> None:
    for col, label in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=label)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = _HEADER_ALIGN


def _autosize(ws, headers: list[str]) -> None:
    """Larghezza colonne basata sull'header (semplice, non scansiona tutti i valori)."""
    for col, label in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(col)].width = max(12, len(label) + 2)


def _write_summary(wb: Workbook, model: FEAModel) -> None:
    ws = wb.create_sheet("Summary", 0)
    rows = [
        ("Field", "Value"),
        ("Model ID", model.id),
        ("Name", model.name or ""),
        ("Description", model.description or ""),
        ("Units", model.units),
        ("Is 3D", "Yes" if model.is_3d else "No"),
        ("Nodes", len(model.nodes)),
        ("Elements", len(model.elements)),
        ("Loads", len(model.loads)),
        ("Constraints", len(model.constraints)),
    ]
    _set_header(ws, 1, list(rows[0]))
    for r, (k, v) in enumerate(rows[1:], start=2):
        ws.cell(row=r, column=1, value=k)
        ws.cell(row=r, column=2, value=v)
    _autosize(ws, list(rows[0]))


def _write_nodes(wb: Workbook, model: FEAModel) -> None:
    ws = wb.create_sheet("Nodes")
    headers = ["id", "x", "y", "z", "label"]
    _set_header(ws, 1, headers)
    for r, n in enumerate(model.nodes, start=2):
        ws.cell(row=r, column=1, value=n.id)
        ws.cell(row=r, column=2, value=n.x)
        ws.cell(row=r, column=3, value=n.y)
        ws.cell(row=r, column=4, value=n.z)
        ws.cell(row=r, column=5, value=n.label or "")
    _autosize(ws, headers)


def _write_elements(wb: Workbook, model: FEAModel) -> None:
    ws = wb.create_sheet("Elements")
    headers = ["id", "type", "nodes", "material_id", "section_id", "winkler_k"]
    _set_header(ws, 1, headers)
    for r, e in enumerate(model.elements, start=2):
        ws.cell(row=r, column=1, value=e.id)
        ws.cell(row=r, column=2, value=e.type.value)
        ws.cell(row=r, column=3, value=", ".join(str(n) for n in e.nodes))
        ws.cell(row=r, column=4, value=e.material_id)
        ws.cell(row=r, column=5, value=e.section_id or "")
        ws.cell(row=r, column=6, value=e.winkler_k)
    _autosize(ws, headers)


def _write_loads(wb: Workbook, model: FEAModel) -> None:
    ws = wb.create_sheet("Loads")
    headers = ["id", "type", "target_id", "fx", "fy", "fz", "qx", "qy", "qz",
               "mx", "my", "mz", "mass", "delta_t", "label"]
    _set_header(ws, 1, headers)
    for r, l in enumerate(model.loads, start=2):
        ws.cell(row=r, column=1, value=l.id)
        ws.cell(row=r, column=2, value=l.type.value)
        ws.cell(row=r, column=3, value=l.target_id)
        ws.cell(row=r, column=4, value=l.fx)
        ws.cell(row=r, column=5, value=l.fy)
        ws.cell(row=r, column=6, value=l.fz)
        ws.cell(row=r, column=7, value=l.qx)
        ws.cell(row=r, column=8, value=l.qy)
        ws.cell(row=r, column=9, value=l.qz)
        ws.cell(row=r, column=10, value=l.mx)
        ws.cell(row=r, column=11, value=l.my)
        ws.cell(row=r, column=12, value=l.mz)
        ws.cell(row=r, column=13, value=l.mass)
        ws.cell(row=r, column=14, value=l.delta_t)
        ws.cell(row=r, column=15, value=l.label or "")
    _autosize(ws, headers)


def _write_constraints(wb: Workbook, model: FEAModel) -> None:
    ws = wb.create_sheet("Constraints")
    headers = ["id", "type", "node_id", "dofs", "spring_k", "compression_only", "label"]
    _set_header(ws, 1, headers)
    for r, c in enumerate(model.constraints, start=2):
        ws.cell(row=r, column=1, value=c.id)
        ws.cell(row=r, column=2, value=c.type.value)
        ws.cell(row=r, column=3, value=c.node_id)
        ws.cell(row=r, column=4, value=str(c.dofs) if c.dofs else "")
        ws.cell(row=r, column=5, value=str(c.spring_k) if c.spring_k else "")
        ws.cell(row=r, column=6, value="Yes" if c.compression_only else "No")
        ws.cell(row=r, column=7, value=c.label or "")
    _autosize(ws, headers)


def _write_static_results(wb: Workbook, results: StaticResults) -> None:
    # Displacements
    ws = wb.create_sheet("Displacements")
    headers = ["node_id", "ux", "uy", "uz", "rx", "ry", "rz"]
    _set_header(ws, 1, headers)
    for r, d in enumerate(results.displacements, start=2):
        ws.cell(row=r, column=1, value=d.node_id)
        ws.cell(row=r, column=2, value=d.ux)
        ws.cell(row=r, column=3, value=d.uy)
        ws.cell(row=r, column=4, value=d.uz)
        ws.cell(row=r, column=5, value=d.rx)
        ws.cell(row=r, column=6, value=d.ry)
        ws.cell(row=r, column=7, value=d.rz)
    _autosize(ws, headers)
    # Reactions
    ws = wb.create_sheet("Reactions")
    headers = ["node_id", "fx", "fy", "fz", "mx", "my", "mz"]
    _set_header(ws, 1, headers)
    for r, R in enumerate(results.reactions, start=2):
        ws.cell(row=r, column=1, value=R.node_id)
        ws.cell(row=r, column=2, value=R.fx)
        ws.cell(row=r, column=3, value=R.fy)
        ws.cell(row=r, column=4, value=R.fz)
        ws.cell(row=r, column=5, value=R.mx)
        ws.cell(row=r, column=6, value=R.my)
        ws.cell(row=r, column=7, value=R.mz)
    _autosize(ws, headers)
    # Element forces
    ws = wb.create_sheet("ElementForces")
    headers = ["element_id",
               "N_i", "Vy_i", "Vz_i", "Mx_i", "My_i", "Mz_i",
               "N_j", "Vy_j", "Vz_j", "Mx_j", "My_j", "Mz_j"]
    _set_header(ws, 1, headers)
    for r, f in enumerate(results.element_forces, start=2):
        ws.cell(row=r, column=1, value=f.element_id)
        ws.cell(row=r, column=2, value=f.N_i)
        ws.cell(row=r, column=3, value=f.Vy_i)
        ws.cell(row=r, column=4, value=f.Vz_i)
        ws.cell(row=r, column=5, value=f.Mx_i)
        ws.cell(row=r, column=6, value=f.My_i)
        ws.cell(row=r, column=7, value=f.Mz_i)
        ws.cell(row=r, column=8, value=f.N_j)
        ws.cell(row=r, column=9, value=f.Vy_j)
        ws.cell(row=r, column=10, value=f.Vz_j)
        ws.cell(row=r, column=11, value=f.Mx_j)
        ws.cell(row=r, column=12, value=f.My_j)
        ws.cell(row=r, column=13, value=f.Mz_j)
    _autosize(ws, headers)


def _write_modal_results(wb: Workbook, results: ModalResults) -> None:
    ws = wb.create_sheet("Modes")
    headers = ["mode", "frequency_Hz", "period_s", "omega_rad_s",
               "participation_x", "participation_y", "participation_z",
               "effective_mass_x", "effective_mass_y", "effective_mass_z"]
    _set_header(ws, 1, headers)
    for r, mode in enumerate(results.modes, start=2):
        ws.cell(row=r, column=1, value=mode.mode)
        ws.cell(row=r, column=2, value=mode.frequency_hz)
        ws.cell(row=r, column=3, value=mode.period)
        ws.cell(row=r, column=4, value=mode.omega)
        ws.cell(row=r, column=5, value=mode.participation_x)
        ws.cell(row=r, column=6, value=mode.participation_y)
        ws.cell(row=r, column=7, value=mode.participation_z)
        ws.cell(row=r, column=8, value=mode.effective_mass_x)
        ws.cell(row=r, column=9, value=mode.effective_mass_y)
        ws.cell(row=r, column=10, value=mode.effective_mass_z)
    _autosize(ws, headers)


def export_excel(
    model: FEAModel,
    file_path: str | Path,
    *,
    static_results: Optional[StaticResults] = None,
    modal_results: Optional[ModalResults] = None,
) -> Path:
    """Esporta modello + opzionalmente risultati in un file .xlsx.

    Args:
        model           : FEAModel
        file_path       : path di output (.xlsx)
        static_results  : se fornito, aggiunge sheet Displacements/Reactions/ElementForces
        modal_results   : se fornito, aggiunge sheet Modes

    Returns:
        Path scritto.
    """
    file_path = Path(file_path)
    wb = Workbook()
    # Rimuovi sheet di default
    if wb.active is not None and "Sheet" in wb.sheetnames:
        del wb["Sheet"]
    _write_summary(wb, model)
    _write_nodes(wb, model)
    _write_elements(wb, model)
    _write_loads(wb, model)
    _write_constraints(wb, model)
    if static_results is not None:
        _write_static_results(wb, static_results)
    if modal_results is not None:
        _write_modal_results(wb, modal_results)
    file_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(file_path))
    return file_path
