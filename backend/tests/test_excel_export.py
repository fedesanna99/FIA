"""
Test export Excel + round-trip openpyxl.

Strategia: scrivi un modello → riapri il workbook → verifica sheet attesi
e valori chiave dei nodi/elementi.
"""
from __future__ import annotations
from pathlib import Path
import io

import pytest
from openpyxl import load_workbook

from schemas import (
    FEAModel, Node, Element, Load, Constraint,
    ElementType, LoadType, ConstraintType,
)
from schemas.results import (
    StaticResults, ModalResults, ModeShape,
    NodalDisplacement, NodalReaction, ElementForces,
)
from core.io import export_excel


def _sample_model() -> FEAModel:
    return FEAModel(
        id="xlsx_t", name="Sample", description="Test export",
        is_3d=False,
        nodes=[Node(id=1, x=0, y=0, z=0, label="A"),
                Node(id=2, x=4, y=0, z=0)],
        elements=[Element(id=1, type=ElementType.BEAM2D, nodes=[1, 2],
                           material_id="steel_s355", section_id="ipe_300",
                           winkler_k=1e5)],
        loads=[Load(id=1, type=LoadType.NODAL, target_id=2, fy=-1000.0, label="P")],
        constraints=[Constraint(id=1, type=ConstraintType.FIXED, node_id=1,
                                 compression_only=False)],
    )


class TestExcelExportBasic:
    def test_creates_file(self, tmp_path: Path):
        p = tmp_path / "out.xlsx"
        out = export_excel(_sample_model(), p)
        assert out == p
        assert p.exists() and p.stat().st_size > 0

    def test_sheets_present(self, tmp_path: Path):
        p = tmp_path / "s.xlsx"
        export_excel(_sample_model(), p)
        wb = load_workbook(p)
        expected = {"Summary", "Nodes", "Elements", "Loads", "Constraints"}
        assert expected.issubset(set(wb.sheetnames))

    def test_nodes_values(self, tmp_path: Path):
        p = tmp_path / "n.xlsx"
        export_excel(_sample_model(), p)
        wb = load_workbook(p)
        ws = wb["Nodes"]
        # Header row
        assert ws.cell(1, 1).value == "id"
        # Riga 1 nodo 1
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == 0
        assert ws.cell(2, 5).value == "A"
        # Riga nodo 2
        assert ws.cell(3, 1).value == 2
        assert ws.cell(3, 2).value == 4

    def test_elements_values(self, tmp_path: Path):
        p = tmp_path / "e.xlsx"
        export_excel(_sample_model(), p)
        wb = load_workbook(p)
        ws = wb["Elements"]
        assert ws.cell(2, 2).value == "beam2d"
        assert ws.cell(2, 3).value == "1, 2"
        assert ws.cell(2, 4).value == "steel_s355"
        assert ws.cell(2, 6).value == 1e5

    def test_loads_values(self, tmp_path: Path):
        p = tmp_path / "l.xlsx"
        export_excel(_sample_model(), p)
        wb = load_workbook(p)
        ws = wb["Loads"]
        assert ws.cell(2, 5).value == -1000.0  # fy
        assert ws.cell(2, 15).value == "P"     # label

    def test_constraints_values(self, tmp_path: Path):
        p = tmp_path / "c.xlsx"
        export_excel(_sample_model(), p)
        wb = load_workbook(p)
        ws = wb["Constraints"]
        assert ws.cell(2, 2).value == "fixed"
        assert ws.cell(2, 3).value == 1


class TestExcelExportWithStaticResults:
    def test_static_sheets_present(self, tmp_path: Path):
        p = tmp_path / "stat.xlsx"
        sr = StaticResults(
            model_id="xlsx_t",
            displacements=[
                NodalDisplacement(node_id=1, ux=0, uy=0, uz=0),
                NodalDisplacement(node_id=2, ux=0.001, uy=-0.02, uz=0),
            ],
            reactions=[
                NodalReaction(node_id=1, fx=0, fy=1000, fz=0,
                              mx=0, my=0, mz=0),
            ],
            element_forces=[
                ElementForces(element_id=1, N_i=0, Vy_i=1000, Mz_i=0,
                              N_j=0, Vy_j=-1000, Mz_j=4000),
            ],
        )
        export_excel(_sample_model(), p, static_results=sr)
        wb = load_workbook(p)
        assert "Displacements" in wb.sheetnames
        assert "Reactions" in wb.sheetnames
        assert "ElementForces" in wb.sheetnames
        # Spot-check
        assert wb["Displacements"].cell(3, 2).value == pytest.approx(0.001)
        # Col 13 = Mz_j
        assert wb["ElementForces"].cell(2, 13).value == pytest.approx(4000)


class TestExcelExportWithModalResults:
    def test_modes_sheet_present(self, tmp_path: Path):
        p = tmp_path / "modal.xlsx"
        mr = ModalResults(
            model_id="xlsx_t", n_modes=2,
            modes=[
                ModeShape(mode=1, frequency_hz=2.5, omega=15.71, period=0.4,
                           participation_x=0.8),
                ModeShape(mode=2, frequency_hz=8.1, omega=50.89, period=0.123,
                           participation_x=0.15),
            ],
        )
        export_excel(_sample_model(), p, modal_results=mr)
        wb = load_workbook(p)
        assert "Modes" in wb.sheetnames
        ws = wb["Modes"]
        assert ws.cell(2, 1).value == 1
        assert ws.cell(2, 2).value == pytest.approx(2.5)
        assert ws.cell(3, 1).value == 2


class TestExcelEndpoint:
    def test_endpoint_returns_xlsx(self):
        from fastapi.testclient import TestClient
        from main import app
        import storage

        m = _sample_model()
        m.id = storage.new_id()
        storage.save_model(m)
        client = TestClient(app)
        try:
            r = client.get(f"/api/io/export/{m.id}/xlsx")
            assert r.status_code == 200, r.text
            # Header content-type contiene spreadsheet
            assert "spreadsheet" in r.headers["content-type"].lower() or \
                   "application" in r.headers["content-type"].lower()
            # Apre il file con openpyxl direttamente da bytes
            wb = load_workbook(io.BytesIO(r.content))
            assert "Summary" in wb.sheetnames
        finally:
            storage.delete_model(m.id)

    def test_endpoint_with_static_results(self):
        from fastapi.testclient import TestClient
        from main import app
        import storage

        # Modello effettivamente solvable
        m = FEAModel(
            id=storage.new_id(), name="solv", is_3d=False,
            nodes=[Node(id=1, x=0, y=0, z=0), Node(id=2, x=3, y=0, z=0)],
            elements=[Element(id=1, type=ElementType.BEAM2D, nodes=[1, 2],
                               material_id="steel_s355", section_id="ipe_300")],
            constraints=[Constraint(id=1, type=ConstraintType.FIXED, node_id=1)],
            loads=[Load(id=1, type=LoadType.NODAL, target_id=2, fy=-1000)],
        )
        storage.save_model(m)
        client = TestClient(app)
        try:
            r = client.get(f"/api/io/export/{m.id}/xlsx?include_static=true")
            assert r.status_code == 200
            wb = load_workbook(io.BytesIO(r.content))
            assert "Displacements" in wb.sheetnames
        finally:
            storage.delete_model(m.id)

    def test_endpoint_404_on_missing(self):
        from fastapi.testclient import TestClient
        from main import app
        client = TestClient(app)
        r = client.get("/api/io/export/no_such_id/xlsx")
        assert r.status_code == 404
